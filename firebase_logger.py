"""
Smart Railway Washroom – Firebase → Google Sheets + Excel Logger
================================================================
Reads MQ135 + DHT22 sensor data from Firebase Realtime DB every 5 seconds
and appends each reading to:
  1. Google Sheets (live, online)
  2. Local Excel file (railway_washroom_gas_log.xlsx)
"""

import time
import datetime
import firebase_admin
from firebase_admin import credentials, db
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook

# ─────────────────────────────────────────────
#  ✏️  CONFIGURE THESE VALUES
# ─────────────────────────────────────────────

FIREBASE_CREDENTIALS_FILE = "serviceAccountKey.json"
FIREBASE_DATABASE_URL      = "https://railwayssystem-bbe8a-default-rtdb.asia-southeast1.firebasedatabase.app/"
FIREBASE_SENSOR_NODE       = "Sensor"   # Root node — reads all fields at once

GOOGLE_SHEETS_CREDENTIALS  = "google_sheets_cred.json"
GOOGLE_SHEET_NAME          = "Railways Washroom Data"
GOOGLE_WORKSHEET_NAME      = "Data"

EXCEL_FILE_PATH            = "railway_washroom_gas_log.xlsx"

POLL_INTERVAL_SECONDS      = 5

# ─────────────────────────────────────────────
#  Gas quality thresholds (ppm)
# ─────────────────────────────────────────────
THRESHOLD_GOOD     = 200
THRESHOLD_MODERATE = 500


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def classify_gas(ppm):
    if ppm < THRESHOLD_GOOD:
        return "Good", "Normal"
    elif ppm < THRESHOLD_MODERATE:
        return "Moderate", "Warning"
    else:
        return "Hazardous", "ALERT"


def init_firebase():
    cred = credentials.Certificate(FIREBASE_CREDENTIALS_FILE)
    firebase_admin.initialize_app(cred, {"databaseURL": FIREBASE_DATABASE_URL})
    print("[Firebase] Connected ✓")


def init_google_sheets():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds  = Credentials.from_service_account_file(GOOGLE_SHEETS_CREDENTIALS, scopes=scopes)
    client = gspread.authorize(creds)
    sheet  = client.open(GOOGLE_SHEET_NAME).worksheet(GOOGLE_WORKSHEET_NAME)
    print(f"[Google Sheets] Connected to '{GOOGLE_SHEET_NAME}' ✓")
    return sheet


def read_all_from_firebase():
    ref = db.reference(FIREBASE_SENSOR_NODE)
    val = ref.get()

    if val is None:
        raise ValueError(f"No data found at Firebase path: {FIREBASE_SENSOR_NODE}")

    if not isinstance(val, dict):
        raise ValueError(f"Unexpected data format from Firebase: {val}")

    # PPM — prefer calculated PPM, fall back to raw ADC
    ppm = val.get("AirQuality_PPM") or val.get("AirQuality_RAW")
    if ppm is None:
        raise ValueError(f"No AirQuality field found. Available keys: {list(val.keys())}")

    temp   = val.get("Temperature")
    hum    = val.get("Humidity")
    status = val.get("Status", "Unknown")

    # Guard against None values from DHT22 read failures
    if temp is None or hum is None:
        raise ValueError("Temperature or Humidity is None — DHT22 may have failed")

    # PIR Sensor
    pir_status = val.get("Motion", "Unknown")
    pir_flag   = int(val.get("MotionFlag", 0))
    
    # Ultrasonic
    distance  = val.get("Distance_cm", -1)
    occupancy = val.get("Occupancy", "Unknown")

    return (float(ppm), float(temp), float(hum), str(status),
            str(pir_status), pir_flag, float(distance), str(occupancy))


def append_to_google_sheets(sheet, timestamp, ppm, temp, hum, quality, alert, esp_status,
                            ir_status, ir_flag, distance, occupancy):
    row = [timestamp, ppm, temp, hum, quality, alert, esp_status,
           ir_status, ir_flag, distance, occupancy]
    sheet.append_row(row, value_input_option="USER_ENTERED")


def append_to_excel(filepath, timestamp, ppm, temp, hum, quality, alert, esp_status,
                    ir_status, ir_flag, distance, occupancy):
    wb = load_workbook(filepath)
    ws = wb["Data"]
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1,  value=timestamp)
    ws.cell(row=next_row, column=2,  value=ppm)
    ws.cell(row=next_row, column=3,  value=temp)
    ws.cell(row=next_row, column=4,  value=hum)
    ws.cell(row=next_row, column=5,  value=quality)
    ws.cell(row=next_row, column=6,  value=alert)
    ws.cell(row=next_row, column=7,  value=esp_status)
    ws.cell(row=next_row, column=8,  value=ir_status)
    ws.cell(row=next_row, column=9,  value=ir_flag)
    ws.cell(row=next_row, column=10, value=distance)
    ws.cell(row=next_row, column=11, value=occupancy)
    wb.save(filepath)


def ensure_headers(sheet):
    headers = ["Timestamp", "PPM", "Temperature (°C)", "Humidity (%)",
               "Air Quality", "Alert Level", "ESP32 Status",
               "PIR Status", "PIR Flag", "Distance (cm)", "Occupancy"]
    if sheet.row_count == 0 or sheet.cell(1, 1).value is None:
        sheet.append_row(headers, value_input_option="USER_ENTERED")
        print("[Google Sheets] Header row written ✓")


# ─────────────────────────────────────────────
#  Main loop
# ─────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  Smart Railway Washroom – Sensor Data Logger")
    print("=" * 55)

    init_firebase()
    sheet = init_google_sheets()
    ensure_headers(sheet)

    print(f"\n[Logger] Starting – polling every {POLL_INTERVAL_SECONDS}s ...\n")
    last_value = None

    while True:
        try:
            ppm, temp, hum, esp_status, ir_status, ir_flag, distance, occupancy = read_all_from_firebase()

            if ppm != last_value:
                timestamp        = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                quality, alert   = classify_gas(ppm)

                append_to_google_sheets(sheet, timestamp, ppm, temp, hum, quality, alert, esp_status,
                                        ir_status, ir_flag, distance, occupancy)
                append_to_excel(EXCEL_FILE_PATH, timestamp, ppm, temp, hum, quality, alert, esp_status,
                                ir_status, ir_flag, distance, occupancy)

                icon = "🟢" if alert == "Normal" else ("🟡" if alert == "Warning" else "🔴")
                print(f"{icon}  {timestamp} | PPM: {ppm:.1f} | Temp: {temp:.1f}°C | Hum: {hum:.1f}% | "
                      f"{quality} | {alert} | IR: {ir_status} | Dist: {distance:.1f}cm | {occupancy} | ESP: {esp_status}")
                last_value = ppm

            else:
                print(f"⏳  {datetime.datetime.now().strftime('%H:%M:%S')} | No change ({ppm:.1f} ppm) — skipping")

        except Exception as e:
            print(f"[ERROR] {e}")

        time.sleep(POLL_INTERVAL_SECONDS)


if __name__ == "__main__":
    main()